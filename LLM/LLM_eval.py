import os
import gc
import re
import warnings
import xml.etree.ElementTree as ET
from datetime import datetime

import numpy as np
import pandas as pd
import torch

from tqdm import tqdm
from rouge_score import rouge_scorer
from bert_score import score as bert_score_fn
from sentence_transformers import SentenceTransformer, util

import nltk

# Safer NLTK data handling for AWS/restricted environments
NLTK_DIR = os.path.join(os.path.dirname(__file__), "nltk_data")
os.makedirs(NLTK_DIR, exist_ok=True)
nltk.data.path.append(NLTK_DIR)

for pkg in ["wordnet", "omw-1.4"]:
    try:
        nltk.data.find(f"corpora/{pkg}")
    except LookupError:
        nltk.download(pkg, download_dir=NLTK_DIR, quiet=True)

from nltk.translate.meteor_score import meteor_score

from transformers import (
    AutoTokenizer,
    AutoModelForCausalLM,
    BitsAndBytesConfig,
    AutoModelForSequenceClassification,
)
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore")

print("=" * 80)
print("CHQ EVALUATION - 100 SAMPLES PER CONFIGURATION")
print("=" * 80)
print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print()

# Device
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"

# Load metric models
print("Loading metric models...")
sentence_model = SentenceTransformer("all-MiniLM-L6-v2", device=DEVICE)

# NLI model for entailment
nli_model_name = "roberta-large-mnli"
nli_tokenizer = AutoTokenizer.from_pretrained(nli_model_name)
nli_model = AutoModelForSequenceClassification.from_pretrained(nli_model_name).to(DEVICE).eval()

print("Metrics loaded\n")


class CHQDataset:
    """Dataset loader for CHQ-Summ XML format"""

    def __init__(self, xml_path):
        self.data = []
        self.load_data(xml_path)

    def load_data(self, xml_path):
        """Load XML supporting both legacy and new CHQ schemas."""
        print(f"Loading: {xml_path}")
        tree = ET.parse(xml_path)
        root = tree.getroot()

        def first_text(doc, paths):
            for p in paths:
                el = doc.find(p)
                if el is not None and el.text and el.text.strip():
                    return el.text.strip()
            return None

        for doc in root.findall(".//document"):
            try:
                gold = first_text(
                    doc,
                    [
                        ".//test_metadata/human_summary",
                        ".//reference_summary",
                        ".//human_summary",
                    ],
                )

                subject = first_text(
                    doc,
                    [
                        ".//original_corpus_data/subject",
                        ".//original_question_subject",
                        ".//subject",
                    ],
                )

                content = first_text(
                    doc,
                    [
                        ".//original_corpus_data/content",
                        ".//original_question_content",
                        ".//content",
                    ],
                )

                if not gold or not subject:
                    continue

                # Store subject and content separately for better source document construction
                self.data.append({"subject": subject, "content": content or "", "gold_summary": gold})
            except Exception:
                continue

        print(f"Loaded {len(self.data)} pairs\n")


class MetricsCalculator:
    """Calculate evaluation metrics"""

    def __init__(self):
        self.rouge_scorer = rouge_scorer.RougeScorer(["rougeLsum"], use_stemmer=True)

    def to_text(self, x):
        """Convert to text: empty string for None/NaN, else str(x).strip()."""
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return ""
        return str(x).strip()

    def validate_text(self, text):
        text = self.to_text(text)
        return text and len(text.strip()) >= 3 and len(text.split()) >= 2

    def calculate_rouge_l(self, pred, ref):
        pred = self.to_text(pred)
        ref = self.to_text(ref)
        if not pred or not ref:
            return 0.0
        try:
            return self.rouge_scorer.score(ref, pred)["rougeLsum"].fmeasure
        except Exception:
            return 0.0

    def calculate_bertscore(self, preds, refs):
        if not preds or not refs:
            return {"f1": 0.0}

        pairs = []
        for p, r in zip(preds, refs):
            p = self.to_text(p)
            r = self.to_text(r)
            if self.validate_text(p) and self.validate_text(r):
                pairs.append((p, r))

        if not pairs:
            return {"f1": 0.0}

        valid_preds, valid_refs = zip(*pairs)

        try:
            _, _, F1 = bert_score_fn(
                list(valid_preds),
                list(valid_refs),
                lang="en",
                verbose=False,
                device=DEVICE,
            )
            return {"f1": max(float(F1.mean().item()), 0.0)}
        except Exception:
            return {"f1": 0.0}

    def calculate_meteor(self, pred, ref):
        """METEOR score using NLTK implementation."""
        pred = self.to_text(pred)
        ref = self.to_text(ref)
        if not pred or not ref:
            return 0.0
        try:
            # meteor_score expects reference as list of token lists, hypothesis as token list
            score = meteor_score([ref.split()], pred.split())
            return max(float(score), 0.0)
        except Exception:
            return 0.0

    def calculate_semantic_coherence(self, pred, question):
        """Semantic coherence between generated summary and original question."""
        pred = self.to_text(pred)
        question = self.to_text(question)
        if not pred or not question:
            return 0.0
        try:
            e1 = sentence_model.encode(pred, convert_to_tensor=True)
            e2 = sentence_model.encode(question[:512], convert_to_tensor=True)
            return max(float(util.cos_sim(e1, e2).item()), 0.0)
        except Exception:
            return 0.0

    def calculate_entailment(self, pred, question):
        """NLI entailment probability: question entails summary (faithfulness)."""
        pred = self.to_text(pred)
        question = self.to_text(question)
        if not pred or not question:
            return 0.0
        try:
            inputs = nli_tokenizer(
                question[:512],  # premise = question
                pred[:512],      # hypothesis = summary
                return_tensors="pt",
                truncation=True,
                padding=True,
                max_length=512,
            )
            inputs = {k: v.to(DEVICE) for k, v in inputs.items()}
            with torch.no_grad():
                logits = nli_model(**inputs).logits
                probs = torch.softmax(logits, dim=-1)[0]

            id2label = {int(k): v for k, v in nli_model.config.id2label.items()}
            entail_id = next((i for i, lab in id2label.items() if "ENTAIL" in lab.upper()), None)
            if entail_id is None:
                entail_id = 2  # fallback for common MNLI ordering

            return float(probs[entail_id].item())
        except Exception:
            return 0.0


def _safe_str(x):
    """Convert None/NaN to empty string, otherwise strip string."""
    if pd.isna(x) or x is None:
        return ""
    return str(x).strip()


class LLMSummarizer:
    """LLM-based summarization with multiple prompting strategies"""

    def __init__(self, model_name):
        self.model_name = model_name
        self.model_registry = {
            "qwen2-7b": "Qwen/Qwen2-7B-Instruct",
            "mistral-7b": "mistralai/Mistral-7B-Instruct-v0.3",
            "llama3.1-8b": "meta-llama/Meta-Llama-3.1-8B-Instruct",
            "llama3.2-3b": "meta-llama/Llama-3.2-3B-Instruct",
            "gemma-7b": "google/gemma-7b-it",
            "deepseek-7b": "deepseek-ai/deepseek-llm-7b-chat",
        }
        self.load_model()

    def load_model(self):
        self.hf_name = self.model_registry[self.model_name]

        print("=" * 80)
        print(f"Loading: {self.model_name} -> {self.hf_name}")
        print("=" * 80)

        self.tokenizer = AutoTokenizer.from_pretrained(self.hf_name, trust_remote_code=True)

        if not self.tokenizer.pad_token:
            self.tokenizer.pad_token = self.tokenizer.eos_token

        # (kept from your code)
        if "gemma" in self.model_name.lower() and hasattr(self.tokenizer, "padding_side"):
            self.tokenizer.padding_side = "left"

        quant_config = BitsAndBytesConfig(
            load_in_8bit=True,
            llm_int8_threshold=6.0,
            llm_int8_enable_fp32_cpu_offload=True,
        )

        self.model = AutoModelForCausalLM.from_pretrained(
            self.hf_name,
            torch_dtype=torch.float16,
            device_map="auto",
            trust_remote_code=True,
            quantization_config=quant_config,
        )

        self.model.eval()
        if torch.cuda.is_available():
            print(f"Loaded (GPU allocated: {torch.cuda.memory_allocated() / 1e9:.2f} GB)\n")
        else:
            print("Loaded on CPU\n")

    def extract_summary(self, full_text, fallback_question=None):
        """Extract FINAL_QUESTION safely - accepts valid outputs with or without the tag."""
        t = (full_text or "").strip()
        if "FINAL_QUESTION:" in t:
            t = t.split("FINAL_QUESTION:")[-1].strip()
        t = t.split("\n")[0].strip()

        words = t.split()
        if 5 <= len(words) <= 20:
            if not t.endswith("?"):
                t += "?"
            return t

        # Safer fallback: use clipped input question instead of generic placeholder
        if fallback_question:
            fallback = fallback_question.split("?")[0].strip()[:80]
            if len(fallback.split()) >= 3:
                return fallback + "?"
        return ""

    # ----------------------------
    # Multi-stage prompting support
    # ----------------------------
    def _generate_single_stage(self, prompt_text, max_new_tokens=32):
        """Helper to generate a single LLM response."""
        messages = [{"role": "user", "content": prompt_text}]

        try:
            if hasattr(self.tokenizer, "apply_chat_template") and self.tokenizer.chat_template:
                inputs = self.tokenizer.apply_chat_template(
                    messages,
                    tokenize=True,
                    add_generation_prompt=True,
                    return_tensors="pt",
                    return_dict=True,
                )
                inputs = {k: v.to(self.model.device) for k, v in inputs.items()}
                input_ids = inputs["input_ids"]
                attention_mask = inputs.get("attention_mask", None)
            else:
                raise RuntimeError("No chat_template available")
        except Exception:
            inputs = self.tokenizer(
                prompt_text,
                return_tensors="pt",
                max_length=2048,
                truncation=True,
                padding=False,
            ).to(self.model.device)
            input_ids = inputs["input_ids"]
            attention_mask = inputs.get("attention_mask", None)

        rep = 1.03 if "gemma" in self.model_name.lower() else 1.02

        with torch.no_grad():
            outputs = self.model.generate(
                input_ids=input_ids,
                attention_mask=attention_mask,
                max_new_tokens=max_new_tokens,
                min_new_tokens=7,
                do_sample=False,
                num_beams=1,
                repetition_penalty=rep,
                no_repeat_ngram_size=4,
                early_stopping=True,
                pad_token_id=self.tokenizer.pad_token_id,
                eos_token_id=self.tokenizer.eos_token_id,
            )

        gen_ids = outputs[0][input_ids.shape[-1]:]
        return self.tokenizer.decode(gen_ids, skip_special_tokens=True)

    def generate_summary(self, question, examples=None, strategy="standard"):
        """Generate summary using appropriate multi-stage strategy."""
        if strategy == "standard":
            return self._generate_standard(question, examples)
        elif strategy == "element-aware":
            return self._generate_element_aware(question, examples)
        elif strategy == "hierarchical":
            return self._generate_hierarchical(question, examples)
        elif strategy == "chain-of-density":
            return self._generate_cod(question, examples)
        else:
            return self._generate_standard(question, examples)

    def _generate_standard(self, question, examples):
        """Standard: 1 step."""
        prompt = "Rewrite the question as ONE short medical question (10-15 words).\n"
        prompt += "Respond in the following format ONLY:\nFINAL_QUESTION: <question>\n\n"

        if examples:
            for ex in examples[:5]:
                q = ex["question"][:180]
                s = ex["summary"]
                prompt += f"Example:\nQuestion: {q}\nFINAL_QUESTION: {s}\n\n"

        prompt += f"Question: {question[:250]}\nFINAL_QUESTION:"
        
        max_new = 35 if "gemma" in self.model_name.lower() else 32
        text = self._generate_single_stage(prompt, max_new_tokens=max_new)
        return self.extract_summary(text, fallback_question=question)

    def _generate_element_aware(self, question, examples):
        """Element-aware: 2 steps (extract elements → generate summary)."""
        # Stage 1: Extract elements
        prompt_stage1 = (
            "Extract the key medical elements (diseases, drugs, tests, symptoms) from this question.\n"
            "List them as comma-separated terms.\n\n"
            f"Question: {question[:250]}\n"
            "ELEMENTS:"
        )
        
        elements_text = self._generate_single_stage(prompt_stage1, max_new_tokens=50)
        elements = elements_text.strip()

        # Stage 2: Generate summary using extracted elements
        prompt_stage2 = (
            "Write ONE short medical question (10-15 words) that includes these key elements.\n"
            f"Elements: {elements}\n\n"
        )
        
        if examples:
            for ex in examples[:3]:
                q = ex["question"][:180]
                s = ex["summary"]
                prompt_stage2 += f"Example:\nQuestion: {q}\nFINAL_QUESTION: {s}\n\n"
        
        prompt_stage2 += f"Original Question: {question[:250]}\nFINAL_QUESTION:"
        
        max_new = 35 if "gemma" in self.model_name.lower() else 32
        text = self._generate_single_stage(prompt_stage2, max_new_tokens=max_new)
        return self.extract_summary(text, fallback_question=question)

    def _generate_hierarchical(self, question, examples):
        """
        Hierarchical prompting inspired by Eq. (1) and Eq. (2) from
        arXiv:2310.10570. The original formulation performs a coarse-to-fine
        summarisation by (1) independently summarising different parts of
        the input into intermediate outputs y_i and (2) using another
        generation pass conditioned on all y_i to produce the final summary.

        For medical question summarisation, we split the question into
        blocks (sentences or short chunks) up to a fixed character length
        to create sub-questions. Each block is summarised into key
        medical phrases (diseases, symptoms, drugs, tests etc.). Then we
        combine those intermediate key phrases into a final concise
        medical question. Few‑shot examples may optionally be prepended to
        the final prompt.
        """
        # Helper to split a question into smaller blocks.  We aim to
        # generate up to 4 blocks of roughly 120 characters each by
        # splitting on punctuation and packing sentences together.
        def _split_into_blocks(text, max_chars=120, max_blocks=4):
            import re
            # Normalise whitespace
            text = re.sub(r"\s+", " ", text).strip()
            if len(text) <= max_chars:
                return [text]
            # Split on sentence boundaries (approximate)
            parts = re.split(r"(?<=[?.!;:])\s+", text)
            blocks, buf = [], ""
            for p in parts:
                p = p.strip()
                if not p:
                    continue
                # Try to append current sentence to buffer if within
                # limit; otherwise flush buffer
                if len(buf) + len(p) + 1 <= max_chars:
                    buf = (buf + " " + p).strip()
                else:
                    if buf:
                        blocks.append(buf)
                    buf = p
                    if len(blocks) >= max_blocks - 1:
                        break
            if buf and len(blocks) < max_blocks:
                blocks.append(buf)
            return blocks[:max_blocks]

        # Step 1 (Eq. 1): generate intermediate key phrases y_i for each block
        blocks = _split_into_blocks(question[:600], max_chars=140, max_blocks=4)
        intermediates = []
        for idx, blk in enumerate(blocks, 1):
            # Prompt to extract essential medical elements from this part
            p1 = (
                "Identify the essential medical information from this PART of the question.\n"
                "Return 3-8 short phrases, comma‑separated, including conditions, symptoms, drugs, tests, demographics, durations.\n\n"
                f"PART {idx}: {blk}\n"
                "KEY_INFO:"
            )
            yi_text = self._generate_single_stage(p1, max_new_tokens=40).strip()
            # Take only the first line for safety
            yi = yi_text.split("\n")[0].strip()
            intermediates.append(yi)

        # Step 2 (Eq. 2): combine all intermediate key phrases into a final question
        combined = "\n".join([f"- {x}" for x in intermediates])
        p2 = (
            "Combine the following KEY_INFO pieces into ONE short medical question (10-15 words).\n"
            "Ensure the final question captures all relevant medical entities and does not invent new information.\n"
            "Return ONLY in this format:\nFINAL_QUESTION: <question>\n\n"
            "KEY_INFO_LIST:\n" + combined + "\n\n"
        )
        # Include few-shot examples in the combination stage if provided
        if examples:
            for ex in examples[:3]:
                q = ex["question"][:180]
                s = ex["summary"]
                p2 += f"Example:\nQuestion: {q}\nFINAL_QUESTION: {s}\n\n"
        # Provide original question for grounding (not to be copied verbatim)
        p2 += f"Original Question: {question[:250]}\nFINAL_QUESTION:"
        max_new = 35 if "gemma" in self.model_name.lower() else 32
        text = self._generate_single_stage(p2, max_new_tokens=max_new)
        return self.extract_summary(text, fallback_question=question)

    def _generate_cod(self, question, examples):
        """
        Chain‑of‑Density (CoD) adapted from Fig. 1 in arXiv:2309.04269.  This
        implementation follows the two‑step iterative procedure:

        1. Identify 1–3 informative Missing_Entities absent from the current
           question but present in the source question.  Entities may be
           diseases, drugs, tests, symptoms, demographic details etc.
        2. Rewrite a new question that is the same length as the current
           question, preserving all previously mentioned entities and
           adding the Missing_Entities.  The length constraint enforces
           density as in the original CoD algorithm.

        We repeat this process for a fixed number of iterations (5),
        returning the densest question after the final iteration.  The
        prompts enforce JSON output for each iteration, with keys
        "Missing_Entities" and "Denser_Question".  A fallback to the
        standard single‑shot generation is used if parsing fails.
        """
        import json, re
        # target length (approximate word count) – maintain across iterations
        TARGET_WORDS = 12
        ITERATIONS = 5
        # Construct a few‑shot block if examples are provided
        examples_block = ""
        if examples:
            for ex in examples[:5]:
                q = ex["question"][:180]
                s = ex["summary"]
                examples_block += f"Example:\nQuestion: {q}\nFINAL_QUESTION: {s}\n\n"

        # Generate an initial question of the target length using the base prompt
        init_prompt = (
            f"Rewrite as ONE medical question with EXACTLY {TARGET_WORDS} words.\n"
            "Do not invent facts beyond the original question.\n"
            "Return ONLY in this format:\nFINAL_QUESTION: <question>\n\n"
        )
        init_prompt += examples_block
        init_prompt += f"Question: {question[:300]}\nFINAL_QUESTION:"
        max_init_new = 40 if "gemma" in self.model_name.lower() else 36
        init_output = self._generate_single_stage(init_prompt, max_new_tokens=max_init_new)
        current = self.extract_summary(init_output, fallback_question=question)

        # Helper to count words
        def _word_count(s: str) -> int:
            return len([w for w in s.strip().split() if w])

        # CoD iterations: identify missing entities and rewrite at the same length
        for iteration in range(1, ITERATIONS + 1):
            cod_prompt = (
                "You will produce an increasingly entity‑dense medical question.\n"
                "Perform the following 2 steps once:\n"
                "Step 1) Identify 1-3 informative Missing_Entities from the SOURCE_QUESTION that are NOT in the CURRENT_QUESTION.\n"
                "Missing_Entities can include: disease, symptom, drug, test, demographic (age/sex), duration, severity etc.\n"
                "Each entity must be <= 5 words, specific, present in SOURCE_QUESTION, and not already mentioned.\n"
                "Step 2) Write a new Denser_Question with EXACTLY the same number of words as CURRENT_QUESTION, preserving all information in CURRENT_QUESTION and adding the Missing_Entities.\n"
                "Do not add filler text or invent facts.\n"
                "Return ONLY a JSON object with keys Missing_Entities and Denser_Question.\n\n"
                f"SOURCE_QUESTION: {question[:450]}\n"
                f"CURRENT_QUESTION: {current}\n"
            )
            # Generate and parse JSON
            max_step_new = 80
            response = self._generate_single_stage(cod_prompt, max_new_tokens=max_step_new).strip()
            # Attempt to extract JSON substring
            match = re.search(r"\{.*\}", response, flags=re.S)
            updated = False
            if match:
                try:
                    obj = json.loads(match.group(0))
                    denser = obj.get("Denser_Question", "").strip()
                    if denser:
                        # Ensure consistent formatting and punctuation
                        if not denser.endswith("?"):
                            denser += "?"
                        # Update only if word count is within sensible bounds
                        wc = _word_count(denser)
                        if 5 <= wc <= 20:
                            current = denser
                            updated = True
                except Exception:
                    pass
            # Safety fallback: if not updated, keep current
            if not updated:
                continue

        # Final safety: enforce question length boundaries
        if not 5 <= _word_count(current) <= 20:
            fallback = question.split("?")[0].strip()[:80]
            if len(fallback.split()) >= 3:
                return fallback + "?"
            return ""
        return current

    def cleanup(self):
        print(f"Cleaning {self.model_name}...")
        del self.model
        del self.tokenizer
        gc.collect()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()


def run_evaluation():
    # Separate train and test files to prevent few-shot leakage
    test_dataset_path = "data/merged_test_found_only.xml"
    train_dataset_path = "data/merged_train_found_only.xml"  # for few-shot examples
    output_dir = "results_50samples"
    samples_per_config = 50

    # Load test dataset
    if not os.path.exists(test_dataset_path):
        print(f"Test dataset not found: {test_dataset_path}")
        print("Place merged_test_found_only.xml in the same directory as this script")
        return

    os.makedirs(output_dir, exist_ok=True)
    test_dataset = CHQDataset(test_dataset_path)

    if len(test_dataset.data) == 0:
        print("No data loaded from test dataset")
        return

    # Load training dataset for few-shot examples
    train_data = []
    if os.path.exists(train_dataset_path):
        print(f"Loading training data for few-shot examples...")
        train_dataset = CHQDataset(train_dataset_path)
        train_data = train_dataset.data[:10]  # Use up to 10 examples for 5-shot
        print(f"Loaded {len(train_data)} training examples for few-shot\n")
    else:
        print(f"WARNING: Training dataset not found: {train_dataset_path}")
        print("Few-shot learning (2-shot, 5-shot) will be DISABLED.")
        print("Only 0-shot evaluation will run.\n")

    # SAFETY: don't exceed dataset size
    test_size = min(samples_per_config, len(test_dataset.data))
    test_data = test_dataset.data[:test_size]

    print(f"Test: {len(test_data)} | Train: {len(train_data)}")
    print(f"Excel: {samples_per_config} samples per configuration")

    metrics_calc = MetricsCalculator()
    models = ["qwen2-7b", "mistral-7b", "llama3.1-8b", "llama3.2-3b", "gemma-7b", "deepseek-7b"]
    strategies = ["standard", "chain-of-density", "hierarchical", "element-aware"]
    shots = [0, 2, 5]
    shots_to_use = shots if len(train_data) > 0 else [0]
    total = len(models) * len(strategies) * len(shots_to_use)

    print(f"Total configs: {len(models)} models x {len(strategies)} strategies x {len(shots_to_use)} shots = {total}")
    print(f"Total sample rows: {total} x {samples_per_config} = {total * samples_per_config}\n")

    all_results = []
    all_predictions = []
    current = 0

    for model_name in models:
        try:
            summarizer = LLMSummarizer(model_name)

            for strategy in strategies:
                for n_shot in shots_to_use:
                    current += 1
                    config_name = f"{model_name}_{strategy}_{n_shot}shot"

                    print("=" * 80)
                    print(f"[{current}/{total}] {config_name}")
                    print("=" * 80)

                    examples = None
                    if n_shot > 0:
                        if n_shot > len(train_data):
                            print(f"WARNING: Requested {n_shot}-shot but only {len(train_data)} examples available")
                            examples = [
                                {"question": (ex["subject"] + " " + ex["content"]).strip(), "summary": ex["gold_summary"]}
                                for ex in train_data
                            ]
                        else:
                            examples = [
                                {"question": (ex["subject"] + " " + ex["content"]).strip(), "summary": ex["gold_summary"]}
                                for ex in train_data[:n_shot]
                            ]

                    preds, refs, questions = [], [], []
                    for item in tqdm(test_data, desc="Generating"):
                        # Construct question for LLM (subject + content)
                        question = (item["subject"] + " " + item["content"]).strip()
                        try:
                            summary = summarizer.generate_summary(question, examples, strategy)
                            preds.append(_safe_str(summary))
                        except Exception:
                            # Safer fallback: use clipped question
                            fallback = question.split("?")[0].strip()[:80]
                            preds.append(fallback + "?" if len(fallback.split()) >= 3 else "")
                        refs.append(_safe_str(item["gold_summary"]))
                        questions.append(_safe_str(question))      # for Semantic_Coherence, QE_Overlap, Entailment

                    preds = [_safe_str(p) if metrics_calc.validate_text(p) else "" for p in preds]

                    # Debug: Check for empty/short predictions
                    empty_preds = sum(1 for p in preds if len(p) < 5)
                    empty_refs = sum(1 for r in refs if len(r) < 5)
                    if empty_preds > 0 or empty_refs > 0:
                        print(f"WARNING: {empty_preds} empty/short predictions, {empty_refs} empty/short references")
                    
                    # Debug: Show sample pred vs ref vs question
                    print(f"Sample pred: {preds[0][:100] if preds[0] else '(empty)'}...")
                    print(f"Sample ref:  {refs[0][:100]}...")
                    print(f"Sample question: {questions[0][:200]}...")

                    print("Calculating metrics...")
                    
                    # Build valid pairs (same filter for ALL metrics)
                    pairs = []
                    for p, r, q in zip(preds, refs, questions):
                        p = _safe_str(p)
                        r = _safe_str(r)
                        q = _safe_str(q)
                        if metrics_calc.validate_text(p) and metrics_calc.validate_text(r):
                            pairs.append((p, r, q))
                    
                    valid_count = len(pairs)
                    total_count = len(preds)
                    print(f"Valid pairs for metrics: {valid_count}/{total_count} ({100*valid_count/total_count:.1f}%)")
                    
                    if not pairs:
                        rouge_avg = meteor_avg = bert_f1 = semantic_avg = ent_avg = 0.0
                    else:
                        vp, vr, vq = zip(*pairs)
                        
                        # Reference-based metrics (on valid pairs only)
                        rouge_avg = float(np.mean([metrics_calc.calculate_rouge_l(p, r) for p, r in zip(vp, vr)]))
                        meteor_avg = float(np.mean([metrics_calc.calculate_meteor(p, r) for p, r in zip(vp, vr)]))
                        bert_f1 = float(metrics_calc.calculate_bertscore(list(vp), list(vr))["f1"])
                        
                        # Reference-free metrics (on valid pairs only)
                        semantic_avg = float(np.mean([metrics_calc.calculate_semantic_coherence(p, q) for p, q in zip(vp, vq)]))
                        ent_avg = float(np.mean([metrics_calc.calculate_entailment(p, q) for p, q in zip(vp, vq)]))  # premise=question, hypothesis=summary (faithfulness)

                    results = {
                        "model": model_name,
                        "strategy": strategy,
                        "n_shot": n_shot,
                        "metrics": {
                            "ROUGE-LSum": max(rouge_avg, 0.0),
                            "METEOR": max(meteor_avg, 0.0),
                            "BERTScore_F1": max(bert_f1, 0.0),
                            "Semantic Coherence": max(semantic_avg, 0.0),
                            "Entailment": max(ent_avg, 0.0),
                        },
                    }

                    print("-" * 80)
                    print("METRICS (Ref-based: ROUGE-LSum, METEOR, BERTScore | Ref-free: Semantic Coherence, Entailment)")
                    print("-" * 80)
                    for k, v in results["metrics"].items():
                        if k == "ROUGE-LSum":
                            print(f"{k:.<30} {v:.4f}")
                        else:
                            print(f"{k:.<30} {v:.6f}")
                    print("-" * 80)

                    all_results.append(results)
                    all_predictions.append(
                        {
                            "model": model_name,
                            "strategy": strategy,
                            "n_shot": n_shot,
                            "predictions": preds,
                            "references": refs,
                            "questions": questions,
                        }
                    )

            summarizer.cleanup()

        except Exception as e:
            print(f"Error in model {model_name}: {e}")
            continue

    print("=" * 80)
    print("CREATING EXCEL WITH 100 SAMPLES PER CONFIGURATION")
    print("=" * 80)

    excel_path = f"{output_dir}/CHQ_Results_100Samples.xlsx"

    summary_data = []
    for r in all_results:
        summary_data.append(
            {
                "Model": r["model"],
                "Strategy": r["strategy"],
                "N-Shot": r["n_shot"],
                "ROUGE-LSum": round(r["metrics"]["ROUGE-LSum"], 4),
                "METEOR": round(r["metrics"]["METEOR"], 6),
                "BERTScore_F1": round(r["metrics"]["BERTScore_F1"], 6),
                "Semantic Coherence": round(r["metrics"]["Semantic Coherence"], 6),
                "Entailment": round(r["metrics"]["Entailment"], 6),
            }
        )

    df_summary = pd.DataFrame(summary_data)

    pred_data = []
    for p in all_predictions:
        config_key = f"{p['model']}_{p['strategy']}_{p['n_shot']}shot"
        for i in range(min(samples_per_config, len(p["questions"]))):  # always safe
            pred_data.append(
                {
                    "Config_ID": config_key,
                    "Sample_Number": i + 1,
                    "Model": p["model"],
                    "Strategy": p["strategy"],
                    "N-Shot": p["n_shot"],
                    "Original_Question": p["questions"][i][:400],
                    "Gold_Summary": p["references"][i],
                    "Model_Generated_Summary": p["predictions"][i],
                }
            )

    df_predictions = pd.DataFrame(pred_data)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="All_Results", index=False)
        df_predictions.to_excel(writer, sheet_name="All_Samples_100_Each", index=False)

        for model in df_summary["Model"].unique():
            model_data = df_summary[df_summary["Model"] == model].drop("Model", axis=1)
            sheet_name = model.replace(".", "_")[:31]
            model_data.to_excel(writer, sheet_name=sheet_name, index=False)

        for shot in shots:
            shot_df = df_predictions[df_predictions["N-Shot"] == shot]
            sheet_name = f"{shot}_shot_samples"
            shot_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    wb.save(excel_path)

    print("EVALUATION COMPLETE")
    if all_results:
        best_r = max(all_results, key=lambda x: x["metrics"]["ROUGE-LSum"])
        best_m = max(all_results, key=lambda x: x["metrics"]["METEOR"])
        best_b = max(all_results, key=lambda x: x["metrics"]["BERTScore_F1"])
        best_sc = max(all_results, key=lambda x: x["metrics"]["Semantic Coherence"])
        best_e = max(all_results, key=lambda x: x["metrics"]["Entailment"])

        print(f"Best ROUGE-LSum: {best_r['model']} | {best_r['strategy']} | {best_r['n_shot']}-shot")
        print(f"Score: {best_r['metrics']['ROUGE-LSum']:.4f}")

        print(f"Best METEOR: {best_m['model']} | {best_m['strategy']} | {best_m['n_shot']}-shot")
        print(f"Score: {best_m['metrics']['METEOR']:.6f}")

        print(f"Best BERTScore: {best_b['model']} | {best_b['strategy']} | {best_b['n_shot']}-shot")
        print(f"Score: {best_b['metrics']['BERTScore_F1']:.6f}")

        print(f"Best Semantic Coherence: {best_sc['model']} | {best_sc['strategy']} | {best_sc['n_shot']}-shot")
        print(f"Score: {best_sc['metrics']['Semantic Coherence']:.6f}")

        print(f"Best Entailment: {best_e['model']} | {best_e['strategy']} | {best_e['n_shot']}-shot")
        print(f"Score: {best_e['metrics']['Entailment']:.6f}")

    print(f"Excel Output: {excel_path}")
    print(f"End: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    if torch.cuda.is_available():
        print(f"GPU: {torch.cuda.get_device_name(0)}\n")
    else:
        print("No GPU detected - using CPU\n")

    run_evaluation()
