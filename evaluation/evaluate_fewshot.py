#!/usr/bin/env python3
"""
All metrics displayed during evaluation
"""

import os
import json
import torch
import xml.etree.ElementTree as ET
from transformers import AutoTokenizer, AutoModelForCausalLM, BitsAndBytesConfig
from rouge_score import rouge_scorer
from bert_score import score as bert_score_fn
from sentence_transformers import SentenceTransformer, util
import numpy as np
import spacy
import warnings
warnings.filterwarnings('ignore')
from summac.model_summac import SummaCZS
from tqdm import tqdm
import gc
from datetime import datetime
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

print("="*80)
print("CHQ SUMMARIZATION EVALUATION ")
print("="*80)
print(f"Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print()

# Initialize metrics models
print("Loading metric models...")
sentence_model = SentenceTransformer('all-MiniLM-L6-v2')
nlp = spacy.load('en_core_web_sm')
summac_model = SummaCZS(granularity='sentence', model_name='vitc')
print("✓ Metrics models loaded\n")

class CHQDataset:
    """Load and process CHQ dataset from XML"""
    def __init__(self, xml_path: str):
        self.data = []
        self.load_data(xml_path)
    
    def load_data(self, xml_path):
        print(f"Loading dataset: {xml_path}")
        tree = ET.parse(xml_path)
        root = tree.getroot()
        docs = root.findall('.//document')
        
        for doc in docs:
            try:
                # Get gold summary
                human_summary = doc.find('.//test_metadata/human_summary')
                if human_summary is None or not human_summary.text:
                    continue
                
                # Get original question
                subject = doc.find('.//original_corpus_data/subject')
                if subject is None or not subject.text:
                    continue
                
                # Combine subject and content for full question
                question = subject.text.strip()
                content = doc.find('.//original_corpus_data/content')
                if content is not None and content.text and content.text.strip():
                    question += ' ' + content.text.strip()
                
                self.data.append({
                    'question': question,
                    'gold_summary': human_summary.text.strip()
                })
            except:
                continue
        
        print(f'✓ Loaded {len(self.data)} question-summary pairs\n')

class MetricsCalculator:
    """Calculate all evaluation metrics with validation"""
    def __init__(self):
        self.rouge_scorer = rouge_scorer.RougeScorer(['rougeL'], use_stemmer=True)
    
    def validate_text(self, text):
        """Validate text is suitable for metrics"""
        if not text or not isinstance(text, str) or len(text.strip()) < 3:
            return False
        return True
    
    def calculate_rouge_l(self, pred, ref):
        """Calculate ROUGE-L score"""
        try:
            if not self.validate_text(pred):
                return 0.0
            scores = self.rouge_scorer.score(ref, pred)
            return scores['rougeL'].fmeasure
        except:
            return 0.0
    
    def calculate_bertscore(self, preds, refs):
        """Calculate BERTScore F1"""
        try:
            # Validate and fix predictions
            valid_preds = []
            for p in preds:
                if self.validate_text(p):
                    valid_preds.append(p)
                else:
                    valid_preds.append("What is the medical condition?")
            
            P, R, F1 = bert_score_fn(
                valid_preds, refs, 
                lang='en', 
                verbose=False, 
                device='cuda' if torch.cuda.is_available() else 'cpu'
            )
            return {'f1': max(F1.mean().item(), 0.0)}
        except Exception as e:
            print(f"BERTScore error: {e}")
            return {'f1': 0.0}
    
    def calculate_semantic_coherence(self, pred, ref):
        """Calculate semantic coherence using sentence embeddings"""
        try:
            if not self.validate_text(pred):
                return 0.0
            emb1 = sentence_model.encode(pred, convert_to_tensor=True)
            emb2 = sentence_model.encode(ref, convert_to_tensor=True)
            score = util.cos_sim(emb1, emb2).item()
            return max(float(score), 0.0)
        except:
            return 0.0
    
    def calculate_summac(self, pred, source):
        """Calculate SummaC score for consistency"""
        try:
            if not self.validate_text(pred):
                return 0.0
            scores = summac_model.score([source], [pred])
            score = scores['scores'][0]
            return float(score) if score > -0.5 else 0.0
        except:
            return 0.0
    
    def extract_entities(self, text):
        """Extract named entities from text"""
        try:
            if not self.validate_text(text):
                return set()
            doc = nlp(text[:1000])  # Limit text length for spacy
            return set([ent.text.lower() for ent in doc.ents])
        except:
            return set()
    
    def calculate_entity_preservation(self, pred, question):
        """Calculate entity preservation score"""
        if not self.validate_text(pred):
            return 0.0
        q_ents = self.extract_entities(question)
        p_ents = self.extract_entities(pred)
        if len(q_ents) == 0:
            return 1.0
        return float(len(q_ents.intersection(p_ents)) / len(q_ents))
    
    def calculate_qe_overlap(self, pred, question):
        """Calculate question-summary word overlap"""
        if not self.validate_text(pred):
            return 0.0
        p_words = set([w.lower() for w in pred.split() if len(w) > 3 and w.isalpha()])
        q_words = set([w.lower() for w in question.split() if len(w) > 3 and w.isalpha()])
        if len(q_words) == 0:
            return 0.0
        return float(len(p_words.intersection(q_words)) / len(q_words))
    
    def calculate_entailment(self, pred, question):
        """Calculate semantic entailment score"""
        try:
            if not self.validate_text(pred):
                return 0.0
            emb1 = sentence_model.encode(pred, convert_to_tensor=True)
            emb2 = sentence_model.encode(question[:512], convert_to_tensor=True)  # Limit length
            score = util.cos_sim(emb1, emb2).item()
            return max(float(score), 0.0)
        except:
            return 0.0

class LLMSummarizer:
    """LLM-based summarization with multiple strategies"""
    def __init__(self, model_name: str):
        self.model_name = model_name
        self.model_registry = {
            'qwen2-7b': 'Qwen/Qwen2-7B-Instruct',
            'mistral-7b': 'mistralai/Mistral-7B-Instruct-v0.3',
            'llama3.1-8b': 'meta-llama/Meta-Llama-3.1-8B-Instruct',
            'llama3.2-3b': 'meta-llama/Llama-3.2-3B-Instruct',
            'gemma-7b': 'google/gemma-7b-it',
            'deepseek-7b': 'deepseek-ai/deepseek-llm-7b-chat'
        }
        self.load_model()
    
    def load_model(self):
        """Load model with 8-bit quantization"""
        hf_name = self.model_registry[self.model_name]
        print(f"\n{'='*80}")
        print(f"Loading model: {self.model_name}")
        print(f"Hugging Face: {hf_name}")
        print(f"{'='*80}")
        
        # Load tokenizer
        self.tokenizer = AutoTokenizer.from_pretrained(
            hf_name, 
            trust_remote_code=True
        )
        
        # Fix tokenizer padding
        if not self.tokenizer.pad_token:
            self.tokenizer.pad_token = self.tokenizer.eos_token
        
        # Special handling for Gemma
        if 'gemma' in self.model_name.lower():
            if hasattr(self.tokenizer, 'padding_side'):
                self.tokenizer.padding_side = 'left'
        
        # Quantization config for memory efficiency
        quant_config = BitsAndBytesConfig(
            load_in_8bit=True,
            llm_int8_threshold=6.0,
            llm_int8_enable_fp32_cpu_offload=True
        )
        
        # Load model
        self.model = AutoModelForCausalLM.from_pretrained(
            hf_name,
            torch_dtype=torch.float16,
            device_map='auto',
            trust_remote_code=True,
            quantization_config=quant_config
        )
        self.model.eval()
        
        if torch.cuda.is_available():
            print(f"✓ Model loaded (GPU memory: {torch.cuda.memory_allocated()/1e9:.2f} GB)")
        else:
            print("✓ Model loaded (CPU mode)")
    
    def create_prompt(self, question, examples, strategy='standard'):
        """Create optimized prompts for each strategy"""
        
        if strategy == 'standard':
            prompt = """Convert to SHORT medical question (10-15 words). COPY the style from examples EXACTLY.
MUST match these patterns:
"What are the [X]?"
"What causes [X]?"
"Why [X]?"
"How [X]?"
"Can [X] cause [Y]?"

"""
            if examples:
                for i, ex in enumerate(examples[:5], 1):
                    q = ex['question'][:160] if len(ex['question']) > 160 else ex['question']
                    prompt += f"{i}. Long: {q}\n   Short: {ex['summary']}\n\n"
            q = question[:200] if len(question) > 200 else question
            prompt += f"Long: {q}\nShort:"
        
        elif strategy == 'chain-of-density':
            prompt = """Find key medical terms, create SHORT question (10-15 words). Match example style.

"""
            if examples:
                for i, ex in enumerate(examples[:5], 1):
                    q = ex['question'][:160] if len(ex['question']) > 160 else ex['question']
                    prompt += f"{i}. Input: {q}\n   Output: {ex['summary']}\n\n"
            q = question[:200] if len(question) > 200 else question
            prompt += f"Input: {q}\nOutput:"
        
        elif strategy == 'hierarchical':
            prompt = """Identify main medical topic, create SHORT question (10-15 words).

"""
            if examples:
                for i, ex in enumerate(examples[:5], 1):
                    q = ex['question'][:160] if len(ex['question']) > 160 else ex['question']
                    prompt += f"{i}. Original: {q}\n   Concise: {ex['summary']}\n\n"
            q = question[:200] if len(question) > 200 else question
            prompt += f"Original: {q}\nConcise:"
        
        elif strategy == 'element-aware':
            prompt = """Keep medical terms, SHORT question (10-15 words).

"""
            if examples:
                for i, ex in enumerate(examples[:5], 1):
                    q = ex['question'][:160] if len(ex['question']) > 160 else ex['question']
                    prompt += f"{i}. Full: {q}\n   Summary: {ex['summary']}\n\n"
            q = question[:200] if len(question) > 200 else question
            prompt += f"Full: {q}\nSummary:"
        
        return prompt
    
    def extract_summary(self, text, prompt, strategy='standard'):
        # Basic extraction - text after prompt
        if len(text) > len(prompt):
            summary = text[len(prompt):].strip()
        else:
            summary = text.strip()
        
        # Strategy-specific keywords
        keywords = {
            'standard': ['Short:', 'Answer:', 'A:'],
            'chain-of-density': ['Output:', 'Out:'],
            'hierarchical': ['Concise:', 'Short:'],
            'element-aware': ['Summary:', 'Sum:']
        }
        
        # Try to extract after keyword
        for kw in keywords.get(strategy, ['Short:']):
            if kw in text:
                parts = text.split(kw)
                if len(parts) > 1:
                    summary = parts[-1].strip()
                    break
        
        # Clean up summary
        summary = summary.split('\n')[0].strip()
        
        # Remove common artifacts
        artifacts = ['Long:', 'Input:', 'Original:', 'Full:', 'Example', 'MUST', 
                    '1.', '2.', '3.', '4.', '5.', '###', 'Note:', 'Task:', '```']
        for art in artifacts:
            if art in summary:
                summary = summary.split(art)[0].strip()
        
        # Clean quotes and formatting
        summary = summary.strip('"').strip("'").strip('`')
        summary = re.sub(r'^\d+[\.\)]\s*', '', summary)
        
        # Limit to 20 words
        words = summary.split()
        if len(words) > 20:
            summary = ' '.join(words[:17])
        
        # If too short, try to find question pattern in text
        if len(words) < 4:
            question_pattern = r'(What|How|Why|Can|Does|Is|Are)\s+[^?.!]{8,80}\?'
            matches = re.findall(question_pattern, text, re.IGNORECASE)
            if matches:
                for match_text in matches:
                    idx = text.find(match_text)
                    if idx >= 0:
                        end_idx = text.find('?', idx) + 1
                        if end_idx > idx:
                            potential = text[idx:end_idx].strip()
                            pot_words = potential.split()
                            if 5 <= len(pot_words) <= 20:
                                summary = potential
                                break
        
        # Ensure question mark for question words
        if summary and not summary.endswith('?'):
            starters = ['what', 'how', 'why', 'when', 'where', 'who', 'is', 'are', 
                       'can', 'does', 'do', 'could', 'should', 'would', 'will']
            if any(summary.lower().startswith(s) for s in starters):
                summary += '?'
        
        # Final validation
        if not summary or len(summary.strip()) < 5 or len(summary.split()) < 3:
            summary = "What is the medical condition?"
        
        return summary.strip()
    
    def generate_summary(self, question, examples=None, strategy='standard'):
        
        prompt = self.create_prompt(question, examples, strategy)
        
        # Special handling for Gemma
        if 'gemma' in self.model_name.lower():
            # Use slightly different settings for Gemma
            max_new_tokens = 35
            temperature = 0.02
            repetition_penalty = 1.03
        else:
            max_new_tokens = 32
            temperature = 0.01
            repetition_penalty = 1.02
        
        # Tokenize input
        inputs = self.tokenizer(
            prompt,
            return_tensors='pt',
            max_length=1600,
            truncation=True,
            padding=True
        ).to(self.model.device)
        
        # Generate with optimized parameters
        with torch.no_grad():
            outputs = self.model.generate(
                **inputs,
                max_new_tokens=max_new_tokens,
                min_new_tokens=7,
                num_beams=15,  # High beam search for quality
                temperature=temperature,  # Ultra-conservative temperature
                top_p=0.8,
                repetition_penalty=repetition_penalty,  # Low penalty to allow phrase copying
                length_penalty=0.4,  # Prefer shorter outputs
                no_repeat_ngram_size=4,
                early_stopping=True,
                do_sample=False,  # Deterministic generation
                pad_token_id=self.tokenizer.pad_token_id,
                eos_token_id=self.tokenizer.eos_token_id
            )
        
        # Decode and extract summary
        full_text = self.tokenizer.decode(outputs[0], skip_special_tokens=True)
        summary = self.extract_summary(full_text, prompt, strategy)
        
        # Final validation
        if not summary or len(summary.strip()) < 5:
            summary = "What is the medical condition?"
        
        return summary
    
    def cleanup(self):
        """Clean up model and free memory"""
        print(f"Cleaning up {self.model_name}...")
        del self.model
        del self.tokenizer
        gc.collect()
        torch.cuda.empty_cache()

def create_excel_output(output_dir, all_results, all_predictions):
    """Create comprehensive Excel with all metrics and samples"""
    
    print("\n" + "="*80)
    print("CREATING EXCEL OUTPUT")
    print("="*80)
    
    excel_path = f'{output_dir}/CHQ_Results_Complete.xlsx'
    
    # Main results with ALL metrics
    summary_data = []
    for r in all_results:
        summary_data.append({
            'Model': r['model'],
            'Strategy': r['strategy'], 
            'N-Shot': r['n_shot'],
            'ROUGE-L': round(r['metrics']['ROUGE-L'], 6),
            'BERTScore_F1': round(r['metrics']['BERTScore_F1'], 6),
            'Semantic_Coherence': round(r['metrics']['Semantic_Coherence'], 6),
            'SummaC': round(r['metrics']['SummaC'], 6),
            'QE_Overlap': round(r['metrics']['QE_Overlap'], 6),
            'Entailment': round(r['metrics']['Entailment'], 6),
            'Entity_Preservation': round(r['metrics']['Entity_Preservation'], 6)
        })
    
    df_summary = pd.DataFrame(summary_data)
    
    # Sample predictions (30 per config)
    pred_data = []
    for pred_info in all_predictions:
        num_samples = min(30, len(pred_info['predictions']))
        for i in range(num_samples):
            pred_data.append({
                'Sample_ID': i + 1,
                'Model': pred_info['model'],
                'Strategy': pred_info['strategy'],
                'N-Shot': pred_info['n_shot'],
                'Original_Question': pred_info['questions'][i][:400],
                'Gold_Summary': pred_info['references'][i],
                'Model_Summary': pred_info['predictions'][i]
            })
    
    df_predictions = pd.DataFrame(pred_data)
    
    # Write Excel with multiple sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # All results sheet
        df_summary.to_excel(writer, sheet_name='All_Results', index=False)
        
        # Per-model sheets
        for model in df_summary['Model'].unique():
            model_data = df_summary[df_summary['Model'] == model].copy()
            model_data = model_data.drop('Model', axis=1)
            sheet_name = model.replace('.', '_')[:31]  # Excel sheet name limit
            model_data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Sample predictions
        df_predictions.to_excel(writer, sheet_name='Sample_Predictions', index=False)
    
    # Format Excel
    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Style header
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    print(f"✅ Main Excel created: {excel_path}")
    return excel_path

def run_evaluation(dataset_path='merged_l6_found_only.xml', output_dir='results_final', test_size=50):
    """Main evaluation pipeline"""
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Load dataset
    dataset = CHQDataset(dataset_path)
    
    if len(dataset.data) == 0:
        print("\n❌ No data loaded!")
        return
    
    # Split data
    test_data = dataset.data[:test_size]
    train_data = dataset.data[test_size:test_size+10] if len(dataset.data) > test_size else dataset.data[:10]
    
    print(f"Test samples: {len(test_data)}")
    print(f"Train examples: {len(train_data)}")
    print(f"🎯 Target: ROUGE-L > 47%, BERTScore > 70%\n")
    
    # Initialize metrics calculator
    metrics_calc = MetricsCalculator()
    
    # Configuration
    models = ['qwen2-7b', 'mistral-7b', 'llama3.1-8b', 'llama3.2-3b', 'gemma-7b', 'deepseek-7b']
    shots = [0, 2, 5]
    strategies = ['standard', 'chain-of-density', 'hierarchical', 'element-aware']
    
    all_results = []
    all_predictions = []
    
    total_configs = len(models) * len(strategies) * len(shots)
    current = 0
    
    for model_name in models:
        try:
            summarizer = LLMSummarizer(model_name)
            
            for strategy in strategies:
                for n_shot in shots:
                    current += 1
                    
                    print(f"\n{'='*80}")
                    print(f"[{current}/{total_configs}] {model_name} | {strategy} | {n_shot}-shot")
                    print(f"{'='*80}\n")
                    
                    # Prepare examples for few-shot
                    examples = None
                    if n_shot > 0:
                        examples = [
                            {'question': ex['question'], 'summary': ex['gold_summary']} 
                            for ex in train_data[:n_shot]
                        ]
                    
                    # Generate summaries
                    preds, refs, questions = [], [], []
                    
                    for item in tqdm(test_data, desc='Generating summaries'):
                        try:
                            summary = summarizer.generate_summary(
                                item['question'], 
                                examples, 
                                strategy
                            )
                            preds.append(summary)
                            refs.append(item['gold_summary'])
                            questions.append(item['question'])
                        except Exception as e:
                            print(f"Error: {e}")
                            preds.append("What is the condition?")
                            refs.append(item['gold_summary'])
                            questions.append(item['question'])
                    
                    # Validate predictions
                    valid_preds = []
                    for p in preds:
                        if metrics_calc.validate_text(p):
                            valid_preds.append(p)
                        else:
                            valid_preds.append("What is the condition?")
                    preds = valid_preds
                    
                    # Show sample predictions for 5-shot
                    if n_shot == 5:
                        print(f"\n{'─'*60}")
                        print("Sample Predictions:")
                        print(f"{'─'*60}")
                        for i in range(min(3, len(preds))):
                            print(f"\n{i+1}. GOLD: {refs[i]}")
                            print(f"   PRED: {preds[i]}")
                    
                    # Calculate ALL metrics
                    print("\nCalculating metrics...")
                    
                    rouge_scores = [metrics_calc.calculate_rouge_l(p, r) for p, r in zip(preds, refs)]
                    bert_scores = metrics_calc.calculate_bertscore(preds, refs)
                    semantic = [metrics_calc.calculate_semantic_coherence(p, r) for p, r in zip(preds, refs)]
                    summac = [metrics_calc.calculate_summac(p, q) for p, q in zip(preds, questions)]
                    qe_overlap = [metrics_calc.calculate_qe_overlap(p, q) for p, q in zip(preds, questions)]
                    entailment = [metrics_calc.calculate_entailment(p, q) for p, q in zip(preds, questions)]
                    entity_pres = [metrics_calc.calculate_entity_preservation(p, q) for p, q in zip(preds, questions)]
                    
                    # Store results
                    results = {
                        'model': model_name,
                        'strategy': strategy,
                        'n_shot': n_shot,
                        'metrics': {
                            'ROUGE-L': max(np.mean(rouge_scores), 0.0),
                            'BERTScore_F1': max(bert_scores['f1'], 0.0),
                            'Semantic_Coherence': max(np.mean(semantic), 0.0),
                            'SummaC': max(np.mean(summac), 0.0),
                            'QE_Overlap': max(np.mean(qe_overlap), 0.0),
                            'Entailment': max(np.mean(entailment), 0.0),
                            'Entity_Preservation': max(np.mean(entity_pres), 0.0)
                        }
                    }
                    
                    # DISPLAY ALL METRICS
                    print(f"\n{'─'*80}")
                    print("RESULTS - ALL METRICS:")
                    print(f"{'─'*80}")
                    print(f"ROUGE-L..................... {results['metrics']['ROUGE-L']:.6f}")
                    print(f"BERTScore_F1................ {results['metrics']['BERTScore_F1']:.6f}")
                    print(f"Semantic_Coherence.......... {results['metrics']['Semantic_Coherence']:.6f}")
                    print(f"SummaC...................... {results['metrics']['SummaC']:.6f}")
                    print(f"QE_Overlap.................. {results['metrics']['QE_Overlap']:.6f}")
                    print(f"Entailment.................. {results['metrics']['Entailment']:.6f}")
                    print(f"Entity_Preservation......... {results['metrics']['Entity_Preservation']:.6f}")
                    print(f"{'─'*80}")
                    
                    # Compare with BART baseline (from paper)
                    bart_rouge = 0.4683
                    bart_bert = 0.6838
                    
                    rouge_diff = (results['metrics']['ROUGE-L'] - bart_rouge) * 100
                    bert_diff = (results['metrics']['BERTScore_F1'] - bart_bert) * 100
                    
                    print("\nComparison with BART baseline:")
                    if rouge_diff > 0:
                        print(f"✅ ROUGE-L: +{rouge_diff:.2f}% vs BART ({bart_rouge:.4f})")
                    else:
                        print(f"❌ ROUGE-L: {rouge_diff:.2f}% vs BART ({bart_rouge:.4f})")
                    
                    if bert_diff > 0:
                        print(f"✅ BERTScore: +{bert_diff:.2f}% vs BART ({bart_bert:.4f})")
                    else:
                        print(f"❌ BERTScore: {bert_diff:.2f}% vs BART ({bart_bert:.4f})")
                    
                    all_results.append(results)
                    all_predictions.append({
                        'model': model_name,
                        'strategy': strategy,
                        'n_shot': n_shot,
                        'predictions': preds,
                        'references': refs,
                        'questions': questions
                    })
            
            # Clean up model
            summarizer.cleanup()
            
        except Exception as e:
            print(f"\n❌ ERROR with {model_name}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    # Create Excel outputs
    excel_path = create_excel_output(output_dir, all_results, all_predictions)
    
    # Final summary
    print("\n" + "="*80)
    print("EVALUATION COMPLETE")
    print("="*80)
    
    if all_results:
        best_rouge = max(all_results, key=lambda x: x['metrics']['ROUGE-L'])
        best_bert = max(all_results, key=lambda x: x['metrics']['BERTScore_F1'])
        
        print(f"\n🏆 Best ROUGE-L: {best_rouge['model']} | {best_rouge['strategy']} | {best_rouge['n_shot']}-shot")
        print(f"   Score: {best_rouge['metrics']['ROUGE-L']:.6f}")
        
        print(f"\n🏆 Best BERTScore: {best_bert['model']} | {best_bert['strategy']} | {best_bert['n_shot']}-shot")
        print(f"   Score: {best_bert['metrics']['BERTScore_F1']:.6f}")
        
        # Summary statistics
        avg_rouge = np.mean([r['metrics']['ROUGE-L'] for r in all_results])
        avg_bert = np.mean([r['metrics']['BERTScore_F1'] for r in all_results])
        
        print(f"\n📊 Average Metrics:")
        print(f"   ROUGE-L: {avg_rouge:.6f}")
        print(f"   BERTScore: {avg_bert:.6f}")
        
        # Count wins vs BART
        rouge_wins = sum(1 for r in all_results if r['metrics']['ROUGE-L'] > 0.4683)
        bert_wins = sum(1 for r in all_results if r['metrics']['BERTScore_F1'] > 0.6838)
        
        print(f"\n📈 Performance vs BART:")
        print(f"   Configurations beating BART ROUGE-L: {rouge_wins}/{len(all_results)}")
        print(f"   Configurations beating BART BERTScore: {bert_wins}/{len(all_results)}")
        
        print(f"\n📁 Output file: {excel_path}")
    
    print(f"\nEnd: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == '__main__':
    # Check dataset exists
    if not os.path.exists('merged_l6_found_only.xml'):
        print("\n❌ Dataset not found! Please upload merged_l6_found_only.xml")
        exit(1)
    
    # Check GPU
    if torch.cuda.is_available():
        print(f"✓ GPU available: {torch.cuda.get_device_name(0)}\n")
    else:
        print("⚠️ No GPU detected - using CPU (will be slower)\n")
    
    # Run evaluation
    run_evaluation(test_size=50)
